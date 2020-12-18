import random
import os
import openpyxl
import scrapy
from scrapy.http import FormRequest
from scrapy import signals

def make_headers():
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{:02d}.0.{:04d}.{} Safari/537.36'.format(
            random.randint(63, 84), random.randint(0, 9999), random.randint(98, 132)),
    }
    return headers

timeout = 100
conn_limit = 200

class MainScraper(scrapy.Spider):
    name = "arrt_scrapy"
    custom_settings = {
        'ROBOTSTXT_OBEY': False,
        'CONCURRENT_REQUESTS': conn_limit,
        # 'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_TARGET_CONCURRENCY': conn_limit,
        # 'AUTOTHROTTLE_START_DELAY ': 1,
        # 'AUTOTHROTTLE_MAX_DELAY ': 360,
        'AUTOTHROTTLE_DEBUG': True,
        # 'DOWNLOAD_DELAY': 1,
        # 'dont_filter': True,
        'RETRY_ENABLED': False,
        # 'COOKIES_ENABLED ': False,
        'CONCURRENT_REQUESTS_PER_DOMAIN': conn_limit,
        'CONCURRENT_REQUESTS_PER_IP': conn_limit,
        'HTTPCACHE_ENABLED': True,
        'HTTPCACHE_IGNORE_HTTP_CODES': [301, 302, 403, 404, 429, 500, 502, 503],
        'HTTPCACHE_STORAGE': 'scrapy.extensions.httpcache.FilesystemCacheStorage',
        'HTTPCACHE_POLICY': 'scrapy.extensions.httpcache.DummyPolicy',
        # 'LOG_ENABLED': False,
        'DOWNLOAD_TIMEOUT': timeout,
        'URLLENGTH_LIMIT': 99999,
    }

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(MainScraper, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_closed, signal=signals.spider_closed)
        return spider

    def __init__(self):
        self.result_dir = os.path.join(os.getcwd(), "Result")
        if not os.path.exists(self.result_dir):
            os.makedirs(self.result_dir)

        self.result_fname = os.path.join(self.result_dir, "Result.csv")
        self.create_result_file()

        heading = [
            "", "", "", "", "", "", "", ""
        ]
        if os.path.getsize(self.result_fname) == 0:
            self.insert_row(result_row=heading)

        self.total_cnt = 0
        self.total_result = []
        self.total_links = []

    def start_requests(self):
        url = ""
        param1 = ""
        request = FormRequest(
            url=url,
            method='GET',
            headers=make_headers(),
            callback=self.get_links,
            errback=self.fail_links,
            dont_filter=True,
            meta={
                'url': url,
                'param1': param1,
                # 'proxy': pxy
                # 'handle_httpstatus_all': True,
                # 'dont_redirect': True,
            }
        )
        yield request

    def get_links(self, response):
        url = response.meta['url']
        param1 = response.meta['param1']
        XPATH = ''

        rows = response.xpath(XPATH)

    def fail_links(self, failure):
        request = FormRequest(
            url=failure.request.meta['url'],
            method='GET',
            headers=make_headers(),
            callback=self.get_links,
            errback=self.fail_links,
            dont_filter=True,
            meta={
                'url': failure.request.meta['url'],
                'param1': failure.request.meta['param1'],
                # 'proxy': pxy
                # 'handle_httpstatus_all': True,
                # 'dont_redirect': True,
            }
        )
        yield request

    # 2. Write xlsx file
    def create_result_file(self):
        if os.path.exists(self.result_fname):
            self.xfile = openpyxl.load_workbook(self.result_fname)
        else:
            self.xfile = openpyxl.Workbook()
        self.sheet = self.xfile.active
        self.row_index = 0

    def insert_row(self, result_row):
        self.row_index += 1
        for i, elm in enumerate(result_row):
            self.sheet.cell(row=self.row_index, column=i + 1).value = elm

    def spider_closed(self, spider):
        self.xfile.save(self.result_fname)

if __name__ == '__main__':
    from scrapy.utils.project import get_project_settings
    from scrapy.crawler import CrawlerProcess

    settings = get_project_settings()
    process = CrawlerProcess(settings)
    process.crawl(MainScraper)
    process.start()